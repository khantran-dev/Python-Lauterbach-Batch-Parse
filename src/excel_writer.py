import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import DebugCable

logger = logging.getLogger(__name__)

# Header columns in exact required order.
# An empty string produces the intentional blank column between
# "QPI Shipped Date" and "Eng Group".
_HEADERS = [
    "Asset tag",
    "Serial number",
    "Serial Number(Secondary)",
    "QC MFG Serial Number",
    "PO Number",
    "Model",
    "State",
    "Substate",
    "Assigned to",
    "Department",
    "Location",
    "Stockroom",
    "Comments",
    "Received Date",
    "Key Contact",
    "Condition",
    "Station name",
    "station type",
    "ECID",
    "Axiom tech team",
    "Memory Upgrade",
    "MSM Upgrade",
    "PMIC Upgrade",
    "RF Upgrade",
    "Local Upgrade Register",
    "WCN Upgrade",
    "IMEI",
    "Cost Center",
    "Parent",
    "Tech team",
    "QPI Shipped Date",
    "",           # intentional blank column
    "Eng Group",
]

_COL_SERIAL = _HEADERS.index("Serial number") + 1
_COL_MODEL  = _HEADERS.index("Model") + 1
_COL_STATE  = _HEADERS.index("State") + 1
_COL_SUBSTATE = _HEADERS.index("Substate") + 1

_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 1-based column indices for required import fields (A=1, B=2, F=6, G=7, H=8, I=9, J=10, K=11, O=15, P=16, AC=29)
_HIGHLIGHT_COLS = {1, 2, 6, 7, 8, 9, 10, 11, 15, 16, 29}

_LICENSE_MODEL_MAP = {
    "LA-2717A": "M049396",
    "LA-2540A": "MDL0063754",
    "LA-2541A": "MDL0063967",
    "LA-3760A": "M025682",
    "LA-3741A": "M016601",
    "LA-3750A": "M058038",
    "LA-3743":  "M056308",
    "LA-7742":  "M007343",
    "LA-7746":  "M007299",
    "LA-7765":  "M007733",
    "LA-7843":  "M008414",
    "LA-7844":  "M014609",
}


class ExcelWriter:
    """Converts a list of DebugCable objects into an .xlsx file."""

    def write(self, cables: list[DebugCable], output_path: Path) -> bool:
        """Write *cables* to *output_path*.  Returns True on success."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Serials"

        # Header row
        ws.append(_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            if cell.column in _HIGHLIGHT_COLS:
                cell.fill = _YELLOW_FILL

        for cable in cables:
            self._write_cable(ws, cable)

        try:
            wb.save(output_path)
            logger.debug("Excel saved to %s", output_path)
            return True
        except OSError as exc:
            logger.error("Failed to save Excel file %s: %s", output_path, exc)
            return False

    # ------------------------------------------------------------------

    def _write_cable(self, ws, cable: DebugCable) -> None:
        # One row for the DebugCable serial itself
        row = _empty_row()
        row[_COL_SERIAL - 1] = cable.serial
        row[_COL_STATE - 1]  = "In use"
        row[_COL_SUBSTATE - 1] = "Issued"
        ws.append(row)

        # One row per license category
        for category in cable.categories:
            row = _empty_row()
            row[_COL_SERIAL - 1]  = category.serial
            row[_COL_STATE - 1]   = "In use"
            row[_COL_SUBSTATE - 1] = "Issued"
            if category.licenses:
                first = category.licenses[0]
                raw_model = f"{first.code} {first.name}"
                row[_COL_MODEL - 1] = _resolve_model_value(raw_model)
            ws.append(row)


def _resolve_model_value(raw_model: str) -> str:
    if not raw_model:
        return raw_model
    license_code = raw_model.split()[0]
    if license_code in _LICENSE_MODEL_MAP:
        return _LICENSE_MODEL_MAP[license_code]
    if license_code.endswith("X"):
        stripped = license_code[:-1]
        if stripped in _LICENSE_MODEL_MAP:
            return _LICENSE_MODEL_MAP[stripped]
    return raw_model


def _empty_row() -> list:
    return [""] * len(_HEADERS)
